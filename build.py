"""
VUTHIES 2026 Dashboard — Build Script

Downloads the data export from Survey Solutions manually, then run:

    python build.py "C:\\Users\\jyaruel\\Downloads\\VUTHIES_2026_1_Tabular_All"

Or just double-click / run without arguments — it will look for the
latest VUTHIES export folder in your Downloads automatically.

After building, push to GitHub Pages with:
    python build.py --push
"""

import csv
import glob
import json
import os
import subprocess
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DOWNLOADS = os.path.join(os.path.expanduser("~"), "Downloads")
QUESTIONNAIRE = "VUTHIES_2026"

# Master lookup file (in Downloads)
HIES_VILLAGES_XLSX = os.path.join(DOWNLOADS, "HIES_Villages.xlsx")

# The .tab files we need
TAB_FILES = {
    f"{QUESTIONNAIRE}.tab": "mainData",
    "hm_basic.tab": "memberData",
    "interview__diagnostics.tab": "diagData",
    "interview__errors.tab": "errorData",
    "interview__comments.tab": "commentData",
    "interview__actions.tab": "actionData",
}


def parse_tsv(filepath):
    """Parse a tab-separated file into a list of dicts."""
    with open(filepath, "r", encoding="utf-8-sig") as f:
        text = f.read()
    lines = text.strip().split("\n")
    if len(lines) < 1:
        return []
    headers = lines[0].split("\t")
    rows = []
    for line in lines[1:]:
        vals = line.split("\t")
        row = {}
        for i, h in enumerate(headers):
            row[h.strip()] = vals[i].strip() if i < len(vals) else ""
        rows.append(row)
    return rows


def find_export_folder():
    """Auto-detect the latest VUTHIES export folder in Downloads."""
    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    pattern = os.path.join(downloads, f"{QUESTIONNAIRE}_*_Tabular_*")
    folders = sorted(glob.glob(pattern), reverse=True)
    for f in folders:
        if os.path.isdir(f) and os.path.exists(os.path.join(f, f"{QUESTIONNAIRE}.tab")):
            return f
    return None


def load_lookups():
    """Load village, area council, and EA lookups from HIES_Villages.xlsx."""
    villages = {}
    area_councils = {}
    ea_lookup = {}

    if os.path.exists(HIES_VILLAGES_XLSX):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(HIES_VILLAGES_XLSX, read_only=True)
            # Use the combined sheet
            if '1&2_Append' in wb.sheetnames:
                ws = wb['1&2_Append']
            else:
                ws = wb.active
            # Headers: VID, Village, Province, Province ID, ACID22, ACNAME22, EA2022, EAHIES
            for row in ws.iter_rows(min_row=2, values_only=True):
                vid, village_name = row[0], row[1]
                acid22, acname22 = row[4], row[5]
                eahies = row[7]
                if vid and village_name:
                    villages[str(int(vid))] = str(village_name)
                if acid22 and acname22:
                    area_councils[str(int(acid22))] = str(acname22)
                if eahies and village_name:
                    ea_key = str(int(eahies))
                    if ea_key not in ea_lookup:
                        ea_lookup[ea_key] = []
                    ea_lookup[ea_key].append(str(village_name))
            wb.close()
            print(f"  Village lookup: {len(villages)} entries")
            print(f"  Area council lookup: {len(area_councils)} entries")
            print(f"  EA lookup: {len(ea_lookup)} EAs")
        except ImportError:
            print("  WARNING: openpyxl not installed — run: pip install openpyxl")
    else:
        print(f"  WARNING: HIES_Villages.xlsx not found at {HIES_VILLAGES_XLSX}")

    return villages, area_councils, ea_lookup


def load_data(src_dir):
    """Read all .tab files from the export folder."""
    data = {}
    for filename, key in TAB_FILES.items():
        filepath = os.path.join(src_dir, filename)
        if os.path.exists(filepath):
            rows = parse_tsv(filepath)
            data[key] = rows
            print(f"  {filename}: {len(rows)} records")
        else:
            print(f"  WARNING: {filename} not found")
            data[key] = []

    # Load and embed lookups
    villages, area_councils, ea_lookup = load_lookups()
    data["villageLookup"] = villages
    data["areaCouncilLookup"] = area_councils
    data["eaLookup"] = ea_lookup

    return data


def build_html(data, src_dir):
    """Inject data into dashboard.html to create self-contained index.html."""
    # Look in script dir first, then in the data export folder
    dashboard_path = os.path.join(SCRIPT_DIR, "dashboard.html")
    if not os.path.exists(dashboard_path):
        dashboard_path = os.path.join(src_dir, "dashboard.html")
    if not os.path.exists(dashboard_path):
        sys.exit(f"ERROR: dashboard.html not found in {SCRIPT_DIR} or {src_dir}")

    with open(dashboard_path, "r", encoding="utf-8") as f:
        html = f.read()

    json_data = json.dumps(data)

    # Find the LOAD DATA FILES marker
    idx_marker = html.find("LOAD DATA FILES")
    if idx_marker < 0:
        sys.exit("ERROR: Could not find 'LOAD DATA FILES' marker in dashboard.html")

    line_start = html.rfind("\n", 0, idx_marker) + 1
    prev_line_start = html.rfind("\n", 0, line_start - 1) + 1

    last_load = html.rfind("loadFile(", idx_marker, idx_marker + 2000)
    end_marker = html.find("// END_REPLACEABLE_BLOCK", last_load)
    if end_marker < 0:
        # Fallback for older dashboard versions
        bracket_end = html.find("]);", last_load)
        end_pos = bracket_end + 3
    else:
        end_pos = end_marker + len("// END_REPLACEABLE_BLOCK")

    replacement = (
        "        // ==========================================================\n"
        "        //  EMBEDDED DATA (built from Survey Solutions export)\n"
        "        // ==========================================================\n"
        "        const EMBEDDED_DATA = " + json_data + ";\n\n"
        "        async function initDashboard() {\n"
        "            document.getElementById('lastUpdate').textContent = 'Loaded: ' + new Date().toLocaleString();\n"
        "\n"
        "            // Use embedded data\n"
        "            const mainData = EMBEDDED_DATA.mainData;\n"
        "            const memberData = EMBEDDED_DATA.memberData;\n"
        "            const diagData = EMBEDDED_DATA.diagData;\n"
        "            const errorData = EMBEDDED_DATA.errorData;\n"
        "            const commentData = EMBEDDED_DATA.commentData;\n"
        "            const actionData = EMBEDDED_DATA.actionData;\n"
        "            const VILLAGE_LOOKUP = EMBEDDED_DATA.villageLookup || {};\n"
        "            const AC_LOOKUP = EMBEDDED_DATA.areaCouncilLookup || {};\n"
        "            const EA_LOOKUP = EMBEDDED_DATA.eaLookup || {};\n"
        "            function villageName(code) { return VILLAGE_LOOKUP[code] || code; }\n"
        "            function acName(code) { return AC_LOOKUP[code] || code; }\n"
        "            function eaVillages(code) { return (EA_LOOKUP[code] || []).join(', ') || code; }\n"
    )

    new_html = html[:prev_line_start] + replacement + html[end_pos:]

    dest = os.path.join(SCRIPT_DIR, "index.html")
    with open(dest, "w", encoding="utf-8") as f:
        f.write(new_html)

    size_kb = os.path.getsize(dest) / 1024
    print(f"\n  Created {dest} ({size_kb:.1f} KB)")
    return dest


def git_push():
    """Commit and push index.html to GitHub."""
    def run(cmd):
        return subprocess.run(cmd, cwd=SCRIPT_DIR, capture_output=True, text=True)

    run(["git", "add", "index.html"])
    result = run(["git", "diff", "--cached", "--quiet"])
    if result.returncode == 0:
        print("  No changes — dashboard is already up to date.")
        return

    import time
    msg = f"Update dashboard data: {time.strftime('%Y-%m-%d %H:%M')}"
    run(["git", "commit", "-m", msg])
    result = run(["git", "push"])
    if result.returncode == 0:
        print("  Pushed to GitHub Pages!")
    else:
        print(f"  Push failed: {result.stderr}")
        print("  You can push manually: git push")


def main():
    print("=" * 50)
    print("VUTHIES 2026 Dashboard Builder")
    print("=" * 50)

    do_push = "--push" in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith("--")]

    # Determine source folder
    if args:
        src_dir = args[0].strip('"').strip("'")
    else:
        src_dir = find_export_folder()
        if not src_dir:
            print("\nERROR: No export folder found.")
            print(f"Download the data from Survey Solutions, then run:")
            print(f'  python build.py "C:\\path\\to\\{QUESTIONNAIRE}_..._Tabular_All"')
            sys.exit(1)

    print(f"\n[1/2] Reading data from:\n  {src_dir}")
    if not os.path.isdir(src_dir):
        sys.exit(f"ERROR: Folder not found: {src_dir}")

    data = load_data(src_dir)

    print(f"\n[2/2] Building dashboard...")
    build_html(data, src_dir)

    if do_push:
        print(f"\n  Pushing to GitHub...")
        git_push()

    print("\nDone!")


if __name__ == "__main__":
    main()
